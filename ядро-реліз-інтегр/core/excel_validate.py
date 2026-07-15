from __future__ import annotations

from pathlib import Path


REQUIRED_HEADERS = {
    "№", "No", "Registration",
    "ІМ'Я", "Ім’я", "Name",
    "Назва статті", "Title",
    "Секція", "Section",
}


def validate_excel_headers(excel_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    headers = {str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)}
    headers_lower = {h.lower() for h in headers}

    def has_any(patterns: set[str]) -> bool:
        return any(any(pat in h for pat in patterns) for h in headers_lower)

    missing = []
    checks = {
        "номер": {"№", "no", "reg", "реєстр", "registration"},
        "ім'я/ПІБ": {"ім", "им", "name", "піб", "автор", "учас", "прізв"},
        "назва": {"назва", "title", "тез"},
        "секція": {"секц", "section"},
    }
    for label, patterns in checks.items():
        if not has_any(patterns):
            missing.append(label)
    if missing:
        details = ", ".join(missing)
        raise ValueError(f"В Excel відсутні обов'язкові колонки: {details}")
