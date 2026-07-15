from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


_HEADER_TITLE = "Назва статті (Excel)"
_HEADER_RELOCATED = "Копія (Статті)"
_HEADER_CLEANED = "Очищена копія"


def _load_matches(run_dir: Path) -> list[dict]:
    matches_path = run_dir / "matches.json"
    if not matches_path.exists():
        return []
    try:
        data = json.loads(matches_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    return data if isinstance(data, list) else []


def _resolve_dashboard_path(run_dir: Path) -> Path | None:
    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            path = manifest.get("dashboard_path")
            if path:
                return Path(path)
        except Exception:
            pass
    # fallback: first xlsx in run_dir
    candidates = list(run_dir.glob("*.xlsx"))
    return candidates[0] if candidates else None


def update_dashboard_paths(
    dashboard_path: Path,
    matches: list[dict],
    *,
    update_relocated: bool,
    update_cleaned: bool,
) -> dict:
    if not dashboard_path.exists():
        return {"updated_relocated": 0, "updated_cleaned": 0, "missing_titles": 0}

    wb = load_workbook(dashboard_path)
    sheet = wb.worksheets[0]

    title_col = None
    relocated_col = None
    cleaned_col = None
    for cell in sheet[1]:
        label = str(cell.value or "").strip()
        if label == _HEADER_TITLE:
            title_col = cell.column
        elif label == _HEADER_RELOCATED:
            relocated_col = cell.column
        elif label == _HEADER_CLEANED:
            cleaned_col = cell.column

    if title_col is None:
        return {"updated_relocated": 0, "updated_cleaned": 0, "missing_titles": 0}

    title_map: dict[str, list[dict]] = {}
    for item in matches:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title") or "").strip()
        if not title:
            continue
        title_map.setdefault(title, []).append(item)

    updated_relocated = 0
    updated_cleaned = 0
    missing_titles = 0

    for row_idx in range(2, sheet.max_row + 1):
        title_cell = sheet.cell(row=row_idx, column=title_col)
        title = str(title_cell.value or "").strip()
        if not title:
            continue
        bucket = title_map.get(title)
        if not bucket:
            missing_titles += 1
            continue
        match = bucket.pop(0)

        if update_relocated and relocated_col is not None:
            relocated = str(match.get("relocated_path") or "")
            if relocated:
                cell = sheet.cell(row=row_idx, column=relocated_col)
                cell.value = relocated
                cell.hyperlink = relocated
                cell.style = "Hyperlink"
                updated_relocated += 1

        if update_cleaned and cleaned_col is not None:
            cleaned = str(match.get("cleaned_path") or "")
            if cleaned:
                cell = sheet.cell(row=row_idx, column=cleaned_col)
                cell.value = cleaned
                cell.hyperlink = cleaned
                cell.style = "Hyperlink"
                updated_cleaned += 1

    wb.save(dashboard_path)
    return {
        "updated_relocated": updated_relocated,
        "updated_cleaned": updated_cleaned,
        "missing_titles": missing_titles,
    }


def update_stats_sheet(
    dashboard_path: Path,
    *,
    relocated_count: int,
    cleaned_count: int,
    stats_sheet_name: str = "Статистика",
) -> None:
    if not dashboard_path.exists():
        return
    wb = load_workbook(dashboard_path)
    if stats_sheet_name in wb.sheetnames:
        stats = wb[stats_sheet_name]
    else:
        stats = wb.create_sheet(stats_sheet_name)

    stats["A11"] = "Оновлення шляхів"
    stats["A11"].font = Font(bold=True)
    stats["A12"] = "relocated_path"
    stats["B12"] = relocated_count
    stats["A13"] = "cleaned_path"
    stats["B13"] = cleaned_count
    stats["A14"] = "updated_at"
    stats["B14"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save(dashboard_path)


def update_dashboard_from_run(
    run_dir: Path,
    *,
    update_relocated: bool,
    update_cleaned: bool,
) -> dict:
    matches = _load_matches(run_dir)
    dashboard_path = _resolve_dashboard_path(run_dir)
    if dashboard_path is None:
        return {"updated_relocated": 0, "updated_cleaned": 0, "missing_titles": 0}
    reopened = False
    excel_visible = False

    try:
        stats = update_dashboard_paths(
            dashboard_path,
            matches,
            update_relocated=update_relocated,
            update_cleaned=update_cleaned,
        )
        update_stats_sheet(
            dashboard_path,
            relocated_count=stats["updated_relocated"],
            cleaned_count=stats["updated_cleaned"],
        )
        return stats
    except PermissionError:
        # Якщо файл відкритий в Excel — закриваємо його, оновлюємо та відкриваємо знову.
        reopened, excel_visible = _close_excel_workbook(dashboard_path)
        stats = update_dashboard_paths(
            dashboard_path,
            matches,
            update_relocated=update_relocated,
            update_cleaned=update_cleaned,
        )
        update_stats_sheet(
            dashboard_path,
            relocated_count=stats["updated_relocated"],
            cleaned_count=stats["updated_cleaned"],
        )
        return stats
    finally:
        if reopened:
            _reopen_excel_workbook(dashboard_path, excel_visible)


def _close_excel_workbook(path: Path) -> tuple[bool, bool]:
    try:
        import pythoncom
        import win32com.client
    except Exception:
        return False, False

    pythoncom.CoInitialize()
    app = None
    try:
        try:
            app = win32com.client.GetObject(None, "Excel.Application")
        except Exception:
            return False, False
        visible = bool(getattr(app, "Visible", False))
        target = str(path.resolve()).casefold()
        for wb in list(app.Workbooks):
            try:
                if str(wb.FullName).casefold() == target:
                    # Зберегти, щоб не втратити зміни користувача.
                    try:
                        wb.Save()
                    except Exception:
                        pass
                    wb.Close(SaveChanges=True)
                    return True, visible
            except Exception:
                continue
        return False, visible
    finally:
        pythoncom.CoUninitialize()


def _reopen_excel_workbook(path: Path, visible: bool) -> None:
    try:
        import pythoncom
        import win32com.client
    except Exception:
        return

    pythoncom.CoInitialize()
    try:
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = bool(visible)
        app.Workbooks.Open(str(path.resolve()))
    except Exception:
        pass
    finally:
        pythoncom.CoUninitialize()
