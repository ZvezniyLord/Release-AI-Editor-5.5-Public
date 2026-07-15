import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")


def add_stats_sheet(path: Path, dashboard_sheet_name: str | None, stats_sheet_name: str) -> None:
    wb = load_workbook(path)
    if dashboard_sheet_name and dashboard_sheet_name in wb.sheetnames:
        dashboard = wb[dashboard_sheet_name]
    else:
        dashboard = wb.worksheets[0]

    if stats_sheet_name in wb.sheetnames:
        wb.remove(wb[stats_sheet_name])

    stats = wb.create_sheet(stats_sheet_name)
    dashboard_name = dashboard.title
    status_col = "H"
    method_col = "I"

    stats["A1"] = "Статистика"
    stats["A1"].font = Font(bold=True, size=14)

    stats["A3"] = "Підсумок"
    stats["A3"].font = Font(bold=True)
    stats["A4"] = "Всього рядків"
    stats["A5"] = "Знайдено"
    stats["A6"] = "Знайдено fallback"
    stats["A7"] = "Не знайдено"
    stats["A8"] = "Відсутній файл статті"
    stats["A9"] = "Вільний слухач"

    stats["B4"] = f"=COUNTA('{dashboard_name}'!{status_col}:{status_col})-1"
    stats["B5"] = f"=COUNTIF('{dashboard_name}'!{status_col}:{status_col},\"Знайдено\")"
    stats["B6"] = f"=COUNTIF('{dashboard_name}'!{status_col}:{status_col},\"Знайдено fallback\")"
    stats["B7"] = f"=COUNTIF('{dashboard_name}'!{status_col}:{status_col},\"Не знайдено\")"
    stats["B8"] = f"=COUNTIF('{dashboard_name}'!{status_col}:{status_col},\"Відсутній файл статті\")"
    stats["B9"] = f"=COUNTIF('{dashboard_name}'!{status_col}:{status_col},\"Вільний слухач\")"

    stats["D3"] = "Методи зіставлення"
    stats["D3"].font = Font(bold=True)
    stats["D4"] = "title_match"
    stats["D5"] = "folder_fallback"
    stats["D6"] = "missing"
    stats["D7"] = "missing_source_file"
    stats["D8"] = "free_listener"

    stats["E4"] = f"=COUNTIF('{dashboard_name}'!{method_col}:{method_col},\"title_match\")"
    stats["E5"] = f"=COUNTIF('{dashboard_name}'!{method_col}:{method_col},\"folder_fallback\")"
    stats["E6"] = f"=COUNTIF('{dashboard_name}'!{method_col}:{method_col},\"missing\")"
    stats["E7"] = f"=COUNTIF('{dashboard_name}'!{method_col}:{method_col},\"missing_source_file\")"
    stats["E8"] = f"=COUNTIF('{dashboard_name}'!{method_col}:{method_col},\"free_listener\")"

    for cell in ("A3", "D3"):
        stats[cell].fill = HEADER_FILL
        stats[cell].font = Font(bold=True)

    stats.column_dimensions["A"].width = 32
    stats.column_dimensions["B"].width = 14
    stats.column_dimensions["D"].width = 26
    stats.column_dimensions["E"].width = 14

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Додає вкладку 'Статистика' у дашборд Excel")
    parser.add_argument("path", type=Path, help="Шлях до dashboard_perevirky.xlsx")
    parser.add_argument("--dashboard-sheet", dest="dashboard_sheet", default=None, help="Назва вкладки дашборду")
    parser.add_argument("--stats-sheet", dest="stats_sheet", default="Статистика", help="Назва вкладки зі статистикою")
    args = parser.parse_args()

    add_stats_sheet(args.path, args.dashboard_sheet, args.stats_sheet)


if __name__ == "__main__":
    main()
