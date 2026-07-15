from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from ..core.models import ArticleRecord
from ..core.sections_loader import load_sections
from ..core.text_utils import canonical

FREE_LISTENER_MARKERS = {
    canonical("Вільний слухач"),
    canonical("Free listener"),
    canonical("Free listeners"),
}


def _normalize_section_key(text: str) -> str:
    value = text or ""
    value = re.sub(r"^\s*(?:секція|section|block)\s*[:№#]?\s*\d+\s*[.)\-:]*\s*", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^\s*\d+\s*[.)\-:]*\s*", "", value)
    key = canonical(value)
    if not key:
        key = canonical(text or "")
    key = re.sub(r"\b(?:та|і|й|и|and)\b", " ", key)
    key = re.sub(r"\s+", " ", key).strip()
    return key


def _compact_section_key(text: str) -> str:
    value = re.sub(r"\s+", "", text or "")
    value = value.replace("та", "")
    return value


def _pick_header(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        expected = canonical(name)
        for header, index in headers.items():
            if canonical(header) == expected:
                return index
    return None


def _extract_section_number(text: str, *, allow_plain_number: bool = False) -> int | None:
    match = re.match(r"^\s*(?:секція|section|block)\s*[:№#]?\s*(\d+)\b", text or "", flags=re.IGNORECASE)
    if match:
        return int(match.group(1))
    if allow_plain_number:
        plain = re.match(r"^\s*(\d+)\s*$", text or "")
        if plain:
            return int(plain.group(1))
    return None


def match_section_label(
    raw_section: str,
    sections: list[dict],
    *,
    allow_plain_number: bool = False,
) -> tuple[int | None, str, str]:
    if not raw_section:
        return None, "", ""
    number = _extract_section_number(raw_section, allow_plain_number=allow_plain_number)
    raw_key = _normalize_section_key(raw_section)
    for section in sections:
        if number is not None and int(section.get("block_number", -1)) == number:
            return number, str(section.get("section_ua", "")), str(section.get("section_en", ""))
    for section in sections:
        ua = _normalize_section_key(str(section.get("section_ua", "")))
        en = _normalize_section_key(str(section.get("section_en", "")))
        raw_compact = _compact_section_key(raw_key)
        ua_compact = _compact_section_key(ua)
        en_compact = _compact_section_key(en)
        if (
            raw_key in {ua, en}
            or raw_key in ua
            or raw_key in en
            or ua in raw_key
            or en in raw_key
            or raw_compact in {ua_compact, en_compact}
            or raw_compact in ua_compact
            or raw_compact in en_compact
            or ua_compact in raw_compact
            or en_compact in raw_compact
        ):
            block_number = section.get("block_number")
            return int(block_number) if block_number else None, str(section.get("section_ua", "")), str(section.get("section_en", ""))
    return None, raw_section, raw_section


def load_articles_from_excel(authors_path: Path, sections_path: Path) -> tuple[list[ArticleRecord], list[dict]]:
    workbook = load_workbook(authors_path, data_only=True)
    worksheet = workbook.active
    headers = {str(worksheet.cell(row=1, column=col).value or "").strip(): col for col in range(1, worksheet.max_column + 1)}

    column_number = _pick_header(headers, "№", "No", "Registration")
    column_author = _pick_header(headers, "ИМЯ", "ІМ'Я", "ІМ’Я", "Name")
    column_title = _pick_header(headers, "Назва статті", "Title")
    column_section = _pick_header(headers, "Секція", "Section")

    sections = load_sections(sections_path)
    free_listener_section_en = next(
        (
            str(section.get("section_en", "")).strip()
            for section in sections
            if str(section.get("block_number")) == "0" and str(section.get("section_en", "")).strip()
        ),
        "Free listeners",
    )
    articles_by_key: dict[str, ArticleRecord] = {}
    raw_rows: list[dict] = []
    free_listener_count = 0

    for row_index in range(2, worksheet.max_row + 1):
        row = {
            str(worksheet.cell(row=1, column=col).value or "").strip(): worksheet.cell(row=row_index, column=col).value
            for col in range(1, worksheet.max_column + 1)
        }
        raw_rows.append(row)

        title = str(worksheet.cell(row=row_index, column=column_title).value or "").strip() if column_title else ""
        author = str(worksheet.cell(row=row_index, column=column_author).value or "").strip() if column_author else ""
        section_raw = str(worksheet.cell(row=row_index, column=column_section).value or "").strip() if column_section else ""
        registration_number = str(worksheet.cell(row=row_index, column=column_number).value or "").strip() if column_number else ""

        if not any((title, author, section_raw)):
            continue

        title_key = canonical(title)
        section_key = canonical(section_raw)
        is_free_listener = not title or title_key in FREE_LISTENER_MARKERS or section_key in FREE_LISTENER_MARKERS
        if is_free_listener:
            free_listener_count += 1
            record = ArticleRecord(
                title=title or "Вільний слухач",
                authors=[author] if author else [],
                section_ua="Вільний слухач",
                section_en=free_listener_section_en,
                section_number=0,
                registration_numbers=[registration_number] if registration_number else [],
                rows=[row],
                is_free_listener=True,
            )
            articles_by_key[f"free_listener_{free_listener_count}"] = record
            continue

        section_number, section_ua, section_en = match_section_label(section_raw, sections)
        key = title_key or f"row_{row_index}"
        record = articles_by_key.get(key)
        if record is None:
            record = ArticleRecord(
                title=title,
                authors=[],
                section_ua=section_ua,
                section_en=section_en,
                section_number=section_number,
            )
            articles_by_key[key] = record
        if author and author not in record.authors:
            record.authors.append(author)
        if registration_number and registration_number not in record.registration_numbers:
            record.registration_numbers.append(registration_number)
        record.rows.append(row)

    articles = list(articles_by_key.values())
    articles.sort(key=lambda item: (item.is_free_listener, item.section_number or 9999, canonical(item.title)))
    return articles, raw_rows
