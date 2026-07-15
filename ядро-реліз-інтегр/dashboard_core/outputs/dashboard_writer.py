from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..core.models import ArticleRecord, MatchRecord

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
GREEN_ALT_FILL = PatternFill(fill_type="solid", start_color="B7E1CD", end_color="B7E1CD")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
BLUE_FILL = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
ORANGE_FILL = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")


def _original_section(article: ArticleRecord) -> str:
    if not article.rows:
        return ""
    row = article.rows[0]
    return str(row.get("Секція") or row.get("Section") or "").strip()


def _original_name(article: ArticleRecord) -> str:
    if article.authors:
        return ", ".join(article.authors)
    if not article.rows:
        return ""
    row = article.rows[0]
    return str(row.get("ИМЯ") or row.get("ІМ'Я") or row.get("Name") or "").strip()


def _original_number(article: ArticleRecord) -> str:
    if article.registration_numbers:
        return str(article.registration_numbers[0])
    if not article.rows:
        return ""
    row = article.rows[0]
    return str(row.get("№") or row.get("No") or row.get("Registration") or "").strip()


def build_dashboard(output_path: Path, articles: list[ArticleRecord], matches: list[MatchRecord]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Дашборд"

    headers = [
        "Ім'я",
        "Назва статті (Excel)",
        "Знайдений файл",
        "Тип",
        "Копія (Статті)",
        "Очищена копія",
        "№",
        "Секція (Excel)",
        "Секція визначена",
        "Номер секції",
        "Статус статті",
        "Метод зіставлення",
        "Конференція",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    matches_by_title = {item.title: item for item in matches}
    sorted_articles = sorted(
        articles,
        key=lambda item: (
            9999 if item.is_free_listener else (item.section_number or 9998),
            item.title.casefold(),
        ),
    )

    for row_index, article in enumerate(sorted_articles, start=1):
        match = matches_by_title.get(article.title)
        if match is not None and match.match_method == "interactive_added":
            status = "Додано вручну"
            fill = YELLOW_FILL
            method = match.match_method
            found_path = match.matched_path
        elif article.is_free_listener:
            status = "Вільний слухач"
            fill = BLUE_FILL
            method = "free_listener"
            found_path = ""
        elif match is not None and match.match_method == "missing_source_file":
            status = "Відсутній файл статті"
            fill = ORANGE_FILL
            method = match.match_method
            found_path = ""
        elif match is None or match.match_method == "missing":
            status = "Не знайдено"
            fill = RED_FILL
            method = match.match_method if match else "missing"
            found_path = ""
        elif "fallback" in match.match_method:
            status = "Знайдено fallback"
            fill = YELLOW_FILL
            method = match.match_method
            found_path = match.matched_path
        else:
            status = "Знайдено"
            fill = GREEN_FILL
            method = match.match_method
            found_path = match.matched_path

        conference = ""
        if article.rows:
            conference = str(article.rows[0].get("Conference") or "").strip()

        cleaned_path = ""
        relocated_path = ""
        assets_flag = ""
        if match is not None:
            cleaned_path = str(getattr(match, "cleaned_path", "") or "")
            relocated_path = str(getattr(match, "relocated_path", "") or "")
            assets = getattr(match, "cleaned_assets", None)
            if assets:
                has_tables = (assets.get("tables", 0) or 0) > 0
                has_media = (assets.get("media_files", 0) or 0) > 0 or (assets.get("embedding_files", 0) or 0) > 0
                if has_tables or has_media:
                    assets_flag = "IMG" if has_media else "TBL"
                else:
                    assets_flag = "TXT"

        row = [
            _original_name(article),
            article.title,
            found_path,
            assets_flag,
            relocated_path,
            cleaned_path,
            _original_number(article),
            _original_section(article),
            match.section_en if match is not None and match.section_en else article.section_en,
            article.section_number if article.section_number is not None else "",
            status,
            method,
            conference,
        ]
        sheet.append(row)
        current_row = sheet.max_row
        row_fill = fill
        if fill == GREEN_FILL:
            row_fill = GREEN_FILL if row_index % 2 == 1 else GREEN_ALT_FILL
        for col_idx in range(1, len(headers) + 1):
            sheet.cell(row=current_row, column=col_idx).fill = row_fill
        file_cell = sheet.cell(row=current_row, column=3)
        if found_path:
            file_cell.hyperlink = found_path
            file_cell.style = "Hyperlink"
        relocated_cell = sheet.cell(row=current_row, column=5)
        if relocated_path:
            relocated_cell.hyperlink = relocated_path
            relocated_cell.style = "Hyperlink"
        cleaned_cell = sheet.cell(row=current_row, column=6)
        if cleaned_path:
            cleaned_cell.hyperlink = cleaned_path
            cleaned_cell.style = "Hyperlink"

    widths = {
        1: 28,
        2: 58,
        3: 14,
        4: 8,
        5: 42,
        6: 42,
        7: 12,
        8: 42,
        9: 36,
        10: 14,
        11: 18,
        12: 22,
        13: 32,
    }
    for col_idx, width in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
